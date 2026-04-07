from __future__ import annotations

import csv, string
import yaml
import requests
from pathlib import Path
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import colors, numbers, PatternFill, Border, Side, Alignment, Protection, Font

from portfolio.handler import make_trade_CptyA, make_trade_CptyB
from portfolio.models_io import (
    InputCheck, BasePosition, BaseTradeCptyA, BaseTradeCptyB, ReportPosition, ReportConcentrationTicker, ReportConcentration,
    ReportReconciliationDetail, ReportReconciliation,
)


def excel_columns():
    for c in string.ascii_uppercase:
        yield c
    for ca in string.ascii_uppercase:
        for cb in string.ascii_uppercase:
            yield ca + cb


def posneg_color(value):
    if value > 0: return "0000ff"
    if value < 0: return "ff0000"
    return "000000"


def read_data_files(data_dir: Path | None = None) -> dict[str, object]:
    """Read the three source files in `data` and return raw content."""
    base_dir = data_dir or (Path(__file__).resolve().parent / "../data")
    files = {
        "cpty_a": base_dir / "cpty_a.csv",
        "cpty_b": base_dir / "cpty_b.csv",
        "position": base_dir / "position.yaml",
    }

    out: dict[str, object] = {}

    with files["cpty_a"].open("r", encoding="utf-8", newline="") as fp:
        out["cpty_a"] = list(csv.DictReader(fp))

    with files["cpty_b"].open("r", encoding="utf-8", newline="") as fp:
        out["cpty_b"] = list(csv.DictReader(fp, delimiter="|"))

    with files["position"].open("r", encoding="utf-8") as fp:
        out["position"] = yaml.safe_load(fp)

    return out


def write_xlsx(trade_date, agg):
    wb = Workbook()

    ws = wb.create_sheet(trade_date.strftime('%Y-%m-%d'), 0)
    ws.active = True
    ws.sheet_properties.tabColor = "1072BA"
    ws.sheet_view.zoomScale = 80
    header = [
        "Account", "Ticker", "Prev Pos Date\n(Empty)", "Prev Pos\n(Empty)", "Curr Pos Date", "Curr Pos",
        "Pos Change\non the Day",
        "Broker A\nQty", "Broker B\nQty", "Total Traded",
        "Discrepancy\nPosition vs Trades", "Discrepancy\nvs API", "API JSON\nResponse",
    ]
    for j, c in enumerate(header):
        cell = ws.cell(row=1, column=j + 1, value=c)
        cell.fill = PatternFill("solid", fgColor="000099")
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.font = Font(bold=True, color=colors.WHITE)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ir = 2
    for key, data in sorted(agg.items()):
        cell = ws.cell(row=ir, column=1, value=key[0])
        cell = ws.cell(row=ir, column=2, value=key[1])
        cell = ws.cell(row=ir, column=5, value=trade_date.strftime('%Y-%m-%d'))

        dpos = data.get('positions', {})
        pos = dpos.get('quantity') or 0
        if pos:
            cell = ws.cell(row=ir, column=6, value=pos)
            cell.font = Font(color=posneg_color(pos))
            cell = ws.cell(row=ir, column=7, value=pos)
            cell.font = Font(color=posneg_color(pos))

        dtrd_a = data.get('trades_a', {})
        qty_a = dtrd_a.get('quantity') or 0
        if qty_a:
            cell = ws.cell(row=ir, column=8, value=qty_a)
            cell.font = Font(color=posneg_color(qty_a))

        dtrd_b = data.get('trades_b', {})
        qty_b = dtrd_b.get('quantity') or 0
        if qty_b:
            cell = ws.cell(row=ir, column=9, value=qty_b)
            cell.font = Font(color=posneg_color(qty_b))

        tot_trd = qty_a + qty_b
        if tot_trd:
            cell = ws.cell(row=ir, column=10, value=tot_trd)
            cell.font = Font(color=posneg_color(tot_trd))

        diff = pos - tot_trd
        if diff:
            cell = ws.cell(row=ir, column=11, value=diff)
            cell.font = Font(color=posneg_color(diff))

        api_diff = data.get('api', {}).get('pos_diff', 0)
        if api_diff:
            cell = ws.cell(row=ir, column=12, value=api_diff)
            cell.font = Font(color=posneg_color(api_diff))

        api_rep = data.get('api', {}).get('json', {})
        if api_rep:
            cell = ws.cell(row=ir, column=13, value=str(api_rep))

        ir += 1
    for j, c in enumerate(excel_columns()):
        if c in ('A', 'B', 'C'): width = 11
        else: width = 10
        ws.column_dimensions[c].width = width
        if j > 20: break
    ws.freeze_panes = ws.cell(row=2, column=1)
    ofn = f'recon-check.{trade_date:%Y-%m-%d}.xlsx'
    wb.save(ofn)


def tabulate_reconciliation() -> None:
    """Tabulate the reconciliation results."""

    payload: dict[str, object] = read_data_files()

    print("Reconciliation results:")
    print(f"cpty_a={len(payload['cpty_a'])}")
    print(f"cpty_b={len(payload['cpty_b'])}")
    print(f"positions={len(payload['position'].get('positions', []))}")

    # A data structure to aggregate and compare quantities from positions and trades
    agg = defaultdict(dict)

    # API request to get the reconciliation results from the server, to compare with our in-memory computation.
    data = requests.get('http://localhost:5000/reconciliation?date=2025-01-15').json()
    check = {}
    for row in data:
        key = ((row['account'], row['ticker']))
        if key in check: raise KeyError(key)
        check[key] = row
        diff = row.get('position', {}).get('quantity', 0) - row.get('trade', {}).get('quantity', 0)
        out = agg[key].setdefault('api', {})
        out.update(pos_diff=diff, json=row)

    # Keys to compare between positions and trades, missing values as 0 for easier comparison and reporting of discrepancies
    compare_on = ['quantity', 'market_value']
    trade_date: datetime = datetime.strptime(payload['position']['report_date'], '%Y%m%d')
    for row in payload['position']['positions']:
        p = BasePosition(trade_date=trade_date, **row)
        key = (p.account, p.ticker)
        out = agg[key].setdefault('positions', {})
        for q in compare_on:
            out[q] = out.get(q, 0) + getattr(p, q)

    for row in payload['cpty_a']:
        t = BaseTradeCptyA(**row)
        t = make_trade_CptyA('CptyA', t, batch_id='A')
        key = (t.account, t.ticker)
        out = agg[key].setdefault('trades_a', {})
        for q in compare_on:
            out[q] = out.get(q, 0) + getattr(t, q)

    for row in payload['cpty_b']:
        t = BaseTradeCptyB(**row)
        t = make_trade_CptyB('CptyB', t, batch_id='B')
        key = (t.account, t.ticker)
        out = agg[key].setdefault('trades_b', {})
        for q in compare_on:
            out[q] = out.get(q, 0) + getattr(t, q)

    write_xlsx(trade_date, agg)

if __name__ == "__main__":
    tabulate_reconciliation()
